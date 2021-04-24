import os
from datetime import date

import openpyxl
from django.db.models import Sum, Count

from capital_repair.models import ContributionsInformation
from dictionaries.models import House
from iggnpk import settings
from tools.date_tools import normalize_date
from tools.get_value import get_value


def to_fixed(numObj, digits=0):
    return f"{numObj:.{digits}f}"


class CrReport:
    _fields = ['assessed_contributions_total',
               'assessed_contributions_current',
               'received_contributions_total',
               'received_contributions_current',
               'delta_total',
               'funds_spent',
               'credit',
               'funds_on_special_deposit',
               'fund_balance']
    _sub_fields = {'total': None,
                   'tsj': ['ТСЖ', 'ТСН'],
                   'jk': ['ЖК', 'ЖСК'],
                   'uk': ['УК', 'ИП'],
                   'ro': ['РО']}
    _report = {}

    mkd_in_reg_program = {}
    total_area = {}
    last_year_funds_spent = {}

    @property
    def assessed_contributions_total(self):
        """Начислено взносов на капитальный ремонт с начала действия региональной программы капитального ремонта
           по начало отчетного периода"""
        result = {}
        for sub_field in self._sub_fields:
            total = self._report['assessed_contributions_total'][sub_field]
            current = self._report['assessed_contributions_current'][sub_field]
            result[sub_field] = (total - current) / 1000000
        return result

    @property
    def received_contributions_total(self):
        """Собрано средств по взносам на капитальный ремонт с начала действия региональной программы
           капитального ремонта по начало отчетного периода"""
        result = {}
        for sub_field in self._sub_fields:
            total = self._report['received_contributions_total'][sub_field]
            current = self._report['received_contributions_current'][sub_field]
            result[sub_field] = (total - current) / 1000000
        return result

    @property
    def level_of_fundraising_total(self):
        """Уровень собираемости средств собственников с начала действия региональной программы капитального ремонта
           по начало отчетного периода"""
        result = {}
        for sub_field in self._sub_fields:
            assessed = self.assessed_contributions_total[sub_field]
            received = self.received_contributions_total[sub_field]
            result[sub_field] = (received * 100) / assessed
        return result

    @property
    def assessed_contributions_current(self):
        """Начислено взносов на капитальный ремонт с начала отчетного периода"""
        result = {}
        for sub_field in self._sub_fields:
            current = self._report['assessed_contributions_current'][sub_field]
            result[sub_field] = current / 1000000
        return result

    @property
    def received_contributions_current(self):
        """Собрано средств по взносам на капитальный ремонт с начала отчетного периода"""
        result = {}
        for sub_field in self._sub_fields:
            current = self._report['received_contributions_current'][sub_field]
            result[sub_field] = current / 1000000
        return result

    @property
    def level_of_fundraising_current(self):
        """Уровень собираемости средств собственников с начала отчетного периода"""
        result = {}
        for sub_field in self._sub_fields:
            assessed = self.assessed_contributions_current[sub_field]
            received = self.received_contributions_current[sub_field]
            result[sub_field] = (received * 100) / assessed
        return result

    @property
    def total_debt(self):
        """Совокупная задолженность собственников по уплате взносов на капитальный ремонт на отчетную дату"""
        result = {}
        for sub_field in self._sub_fields:
            assessed = self.assessed_contributions_total[sub_field]
            fund_balance = self._report['fund_balance'][sub_field]
            last_year_funds_spent = self.last_year_funds_spent[sub_field]
            result[sub_field] = (assessed - fund_balance - last_year_funds_spent) / 1000000
        return result

    @property
    def fund_balance_total(self):
        """Остаток денежных средств на начало отчетного периода"""
        result = {}
        for sub_field in self._sub_fields:
            received = self.received_contributions_current[sub_field]
            fund_balance = self._report['fund_balance'][sub_field]
            last_year_funds_spent = self.last_year_funds_spent[sub_field]
            result[sub_field] = (fund_balance - received + last_year_funds_spent) / 1000000
        return result

    @property
    def fund_balance_total_and_current(self):
        """Остаток денежных средств на начало отчетного периода"""
        result = {}
        for sub_field in self._sub_fields:
            fund_balance = self._report['fund_balance'][sub_field]
            result[sub_field] = fund_balance / 1000000
        return result

    def _initiate_report(self):
        for field in self._fields:
            self._report[field] = {}
        for field in self._fields:
            for item_field in self._sub_fields:
                self._report[field][item_field] = 0

    def _populate_report_dict(self, query):
        for field in self._fields:
            for sub_field, value in self._sub_fields.items():
                if value:
                    self._report[field][sub_field] = query.filter(notify__organization__type__text__in=value) \
                        .aggregate(Sum(field))[field + '__sum']
                else:
                    self._report[field][sub_field] = query.aggregate(Sum(field))[field + '__sum']

    def _populate_reg_program_dict(self, query):
        # количество и площадь домов в рег программе
        for sub_field, value in self._sub_fields.items():
            if value:
                q = House.objects.filter(organization__type__text__in=value, included_in_the_regional_program=True). \
                    aggregate(Sum('total_area'), Count('id'))
            else:
                q = House.objects.filter(included_in_the_regional_program=True). \
                    aggregate(Sum('total_area'), Count('id'))
            self.mkd_in_reg_program[sub_field] = q['id' + '__count']
            if self.mkd_in_reg_program[sub_field] > 0:
                self.total_area[sub_field] = to_fixed((q['total_area' + '__sum'] / 1000), 2)
            else:
                self.total_area[sub_field] = 0

    def _populate_last_year_funds_spent_dict(self, query):
        for sub_field, value in self._sub_fields.items():
            if value:
                self.last_year_funds_spent[sub_field] = query.filter(notify__organization__type__text__in=value) \
                                                            .aggregate(Sum('funds_spent'))['funds_spent__sum'] or 0
            else:
                self.last_year_funds_spent[sub_field] = query.aggregate(Sum('funds_spent'))['funds_spent__sum'] or 0

    def __init__(self, date_start, date_end):
        self._initiate_report()
        date_start = normalize_date(date_start)
        date_end = normalize_date(date_end)
        query = ContributionsInformation.objects.filter(date__range=(date_start, date_end), status_id=3)
        self._populate_report_dict(query)
        self._populate_reg_program_dict(query)
        # инфа по потраченым средствам в прошлом году
        date_start = date(date.today().year - 1, 1, 1)
        date_end = date(date.today().year - 1, 12, 31)
        query = ContributionsInformation.objects.filter(date__range=(date_start, date_end), status_id=3)
        self._populate_last_year_funds_spent_dict(query)




def report_to_excel(report):
    wb = openpyxl.load_workbook(os.path.join(settings.MEDIA_ROOT, 'templates', 'cr_report.xlsx'))
    ws = wb.worksheets[0]
    a = report.assessed_contributions_total
    b = report.assessed_contributions_current
    c = report.received_contributions_total
    d = report.received_contributions_current
    e = report.delta_total
    f = report.funds_spent
    g = report.credit
    h = report.funds_on_special_deposit
    i = report.fund_balance

    j = a - b
    k = c - d
    l = f  # тут должны быть данные за прошлый год

    ws.cell(22, 7).value = j.total / 1000000
    ws.cell(24, 7).value = j.ro / 1000000
    ws.cell(25, 7).value = j.tsj / 1000000
    ws.cell(26, 7).value = j.jk / 1000000
    ws.cell(27, 7).value = j.uk / 1000000

    ws.cell(28, 7).value = k.total / 1000000
    ws.cell(30, 7).value = k.ro / 1000000
    ws.cell(31, 7).value = k.tsj / 1000000
    ws.cell(32, 7).value = k.jk / 1000000
    ws.cell(33, 7).value = k.uk / 1000000

    ws.cell(34, 7).value = (k.total * 100) / j.total
    ws.cell(36, 7).value = (k.ro * 100) / j.ro
    ws.cell(37, 7).value = (k.tsj * 100) / j.tsj
    ws.cell(38, 7).value = (k.jk * 100) / j.jk
    ws.cell(39, 7).value = (k.uk * 100) / j.uk

    ws.cell(40, 7).value = b.total / 1000000
    ws.cell(42, 7).value = b.ro / 1000000
    ws.cell(43, 7).value = b.tsj / 1000000
    ws.cell(44, 7).value = b.jk / 1000000
    ws.cell(45, 7).value = b.uk / 1000000

    ws.cell(46, 7).value = d.total / 1000000
    ws.cell(48, 7).value = d.ro / 1000000
    ws.cell(49, 7).value = d.tsj / 1000000
    ws.cell(50, 7).value = d.jk / 1000000
    ws.cell(51, 7).value = d.uk / 1000000

    ws.cell(52, 7).value = (d.total * 100) / b.total
    ws.cell(54, 7).value = (d.ro * 100) / b.ro
    ws.cell(55, 7).value = (d.tsj * 100) / b.tsj
    ws.cell(56, 7).value = (d.jk * 100) / b.jk
    ws.cell(57, 7).value = (d.uk * 100) / b.uk

    ws.cell(59, 7).value = (a.total - i.total - l.total) / 1000000
    ws.cell(61, 7).value = (a.ro - i.ro - l.ro) / 1000000
    ws.cell(62, 7).value = (a.tsj - i.tsj - l.tsj) / 1000000
    ws.cell(63, 7).value = (a.jk - i.jk - l.jk) / 1000000
    ws.cell(64, 7).value = (a.uk - i.uk - l.uk) / 1000000

    ws.cell(66, 7).value = (i.total - d.total + l.total) / 1000000
    ws.cell(67, 7).value = (i.ro - d.ro + l.ro) / 1000000
    ws.cell(68, 7).value = (i.tsj - d.tsj + l.tsj) / 1000000
    ws.cell(69, 7).value = (i.jk - d.jk + l.jk) / 1000000
    ws.cell(70, 7).value = (i.uk - d.uk + l.uk) / 1000000

    ws.cell(71, 7).value = d.total / 1000000
    ws.cell(72, 7).value = d.ro / 1000000
    ws.cell(73, 7).value = d.tsj / 1000000
    ws.cell(74, 7).value = d.jk / 1000000
    ws.cell(75, 7).value = d.uk / 1000000

    ws.cell(76, 7).value = l.total / 1000000
    ws.cell(77, 7).value = l.ro / 1000000
    ws.cell(78, 7).value = l.tsj / 1000000
    ws.cell(79, 7).value = l.jk / 1000000
    ws.cell(80, 7).value = l.uk / 1000000

    ws.cell(81, 7).value = 0
    ws.cell(82, 7).value = 0
    ws.cell(83, 7).value = 0
    ws.cell(84, 7).value = 0
    ws.cell(85, 7).value = 0

    ws.cell(86, 7).value = i.total / 1000000
    ws.cell(87, 7).value = i.ro / 1000000
    ws.cell(88, 7).value = i.tsj / 1000000
    ws.cell(89, 7).value = i.jk / 1000000
    ws.cell(90, 7).value = i.uk / 1000000

    wb.save(os.path.join(settings.MEDIA_ROOT, 'temp', 'cr_report.xlsx'))
