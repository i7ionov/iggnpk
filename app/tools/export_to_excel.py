from datetime import datetime
import openpyxl
from django.core.exceptions import FieldDoesNotExist
from django.db import models
from django.core.mail import EmailMessage
from openpyxl.writer.excel import save_virtual_workbook

from tools.get_value import get_value


def export_to_excel(template_path, query_set, email):
    """ Экспортирует записи модели в формате xlsx и отправляет на указанный email.
        :param template_path путь к шаблону excel
        :param query_set query set для экспорта"""
    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]
    fields = []
    row_offset = 5
    col = 0
    while ws.cell(row_offset, col + 1).value:
        fields.append(ws.cell(row_offset, col + 1).value)
        col += 1
    for row, item in enumerate(query_set):
        for col, field in enumerate(fields):
            try:
                field_class = item._meta.model._meta.get_field(field.split('.')[0]).__class__
            except FieldDoesNotExist:
                field_class = None
            if field_class == models.fields.related.ManyToManyField:
                # для полей типа ManyToMany просто перечисляем через запятую все найденные значения
                # это значит, что в cвязанной модели должна быть прописана функция __str__
                val = ''
                f = item
                for p in str(field).split('.'):
                    if f is None:
                        continue
                    f = f.__getattribute__(p)
                if f is None:
                    continue
                for i in f.all():
                    if val == '':
                        val = str(i)
                    else:
                        val = val + ';' + str(i)
                ws.cell(row + row_offset, col + 1).value = val
            elif field_class == models.fields.DateField:
                val = get_value(item, field)
                if val:
                    val = datetime.strptime(get_value(item, field),
                                            '%d.%m.%Y')
                    ws.cell(row + row_offset, col + 1).value = val
                    ws.cell(row + row_offset, col + 1).number_format = 'dd.mm.YYYY'
            else:
                # с этого момента начинаем брать каскадом значения полей
                # например field['prefix'] может быть равен 'organization.org_type.'
                # а field['name'] равен 'text'
                # в таком случае наша задача получить значение поля item.organization.org_type.text
                val = get_value(item, field)
                ws.cell(row + row_offset, col + 1).value = val

    file = save_virtual_workbook(wb)
    # wb.save('text.xlsx') # для теста
    email = EmailMessage(
        f'Выгрузка в iggnpk.ru',
        '',
        'noreply@iggnpk.ru',
        [email],
        headers={'Reply-To': 'noreply@iggnpk.ru'}
    )
    email.attach('export.xlsx', file, 'application/xlsx')
    email.send()

    return 'Hello there!'


def get_model_columns(field_list, model, prefix='', parent_verbose_name=''):
    """
    Метод возвращает список полей модели у которых было заполнено verbose_name, работает рекурсивно.
    :param field_list: Объект с полями: verbose_name, name, prefix, field
    :param model: Модель, поля которой нужно получить
    :param prefix: Префикс для поля. Используется в итерациях рекурсии. Хранит имя родительского поля
    :param parent_verbose_name: Используется в итерациях рекурсии. Хранит человекочитаемое название
    родительского поля
    :return: Возвращает список объектов с полями:
    verbose_name - человекочитаемое название поля,
    name - имя поля,
    prefix - имя родительского поля с точкой на конце,
    field - ссылка на сам объект поля
    """
    # выберем поля с verbose_name
    for col, field in enumerate(model._meta.get_fields()):
        field_verbose_name = ''
        if hasattr(field, 'verbose_name') and field.verbose_name:
            try:
                if field.__class__ == models.fields.AutoField or \
                        field.__class__ == models.OneToOneField:
                    continue
                if field.__class__ == models.fields.related.ForeignKey or \
                        field.__class__ == models.fields.related.ManyToManyRel:
                    field_list = get_model_columns(field_list, field.related_model, prefix + field.name + '.',
                                                   field.verbose_name)
                else:
                    if field.name == 'text':
                        field_verbose_name = parent_verbose_name
                    else:
                        field_verbose_name = field.verbose_name
                    field_list.append(
                        {'verbose_name': field_verbose_name, 'name': field.name, 'prefix': prefix, 'field': field})
            except AttributeError:
                pass
    return field_list
